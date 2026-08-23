# -*- coding: utf-8 -*-
import re
from collections import Counter
from pathlib import Path

import openpyxl

p = next(Path(r"C:\Users\admin\FPL_LLM\docs\survey-results").glob("379*.xlsx"))
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb.active
q7 = []
for r in range(2, ws.max_row + 1):
    s2 = ws.cell(r, 9).value
    if s2 not in (1, 2, 3):
        continue
    v = ws.cell(r, 43).value
    if not v:
        continue
    t = str(v).strip()
    if t in ("(空)", "空", "无", "没有", "-", "n/a", "N/A") or not t:
        continue
    q7.append(t)

Path(r"C:\Users\admin\FPL_LLM\docs\survey-results\_q7_utf8.txt").write_text(
    "\n".join(q7), encoding="utf-8"
)

patterns = [
    (r"fantasyfootballscout|fantasy football scout|\bffs\b|ffscout", "Fantasy Football Scout / FFS"),
    (r"fpl focal|fplfocal", "FPL Focal"),
    (r"fpl harry|fplharry|\bharry\b", "FPL Harry"),
    (r"all about fantasy|allaboutfantasy|allaboutfpl", "All About Fantasy"),
    (r"faleague", "FALEAGUE"),
    (r"bigman", "Big Man FPL"),
    (r"\bandy\b", "Andy (creator)"),
    (r"sam fpl|\bsam\b", "Sam / Sam FPL"),
    (r"fplgameweek|fpl gameweek", "FPL Gameweek"),
    (r"let.?s talk", "Let's Talk FPL"),
    (r"reddit", "Reddit"),
]

c = Counter()
for t in q7:
    low = t.lower()
    for pat, lab in patterns:
        if re.search(pat, low, re.I):
            c[lab] += 1

lines = [f"n_answered={len(q7)}"]
for lab, n in c.most_common():
    pct = round(100 * n / len(q7), 1)
    lines.append(f"{lab}\t{n}\t{pct}% of answered")
Path(r"C:\Users\admin\FPL_LLM\docs\survey-results\_q7_brands.txt").write_text(
    "\n".join(lines), encoding="utf-8"
)
print("\n".join(lines))
