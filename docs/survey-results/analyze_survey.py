# -*- coding: utf-8 -*-
"""Analyze Wenjuanxing FPL survey export → deck insights JSON + markdown."""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(r"C:\Users\admin\FPL_LLM\docs\survey-results")
xlsx = next(ROOT.glob("*.xlsx"))
OUT_JSON = ROOT / "analysis-summary.json"
OUT_MD = ROOT / "deck-insights-v2.md"

# Option maps (1-indexed as Wenjuanxing stores for single-choice)
S1 = {1: "是，本赛季", 2: "是，但本赛季还没开始", 3: "以前玩过", 4: "否"}
S2 = {
    1: "中国大陆",
    2: "港澳台",
    3: "新马",
    4: "英国",
    5: "欧洲其他",
    6: "北美",
    7: "印度南亚",
    8: "中东非洲",
    9: "其他",
}
S3 = {1: "仅英文", 2: "仅中文", 3: "中英双语", 4: "其他"}
Q1 = {1: "首季", 2: "2–3", 3: "4–6", 4: "7+"}
Q2 = {1: "前1万", 2: "前10万", 3: "前100万", 4: "100万以外", 5: "不愿说"}
Q4 = {1: "<1小时", 2: "1–3小时", 3: "3–5小时", 4: "5小时以上"}
Q9 = {
    1: "重度依赖",
    2: "需要翻译摘要才用",
    3: "语言障碍很少用",
    4: "不用",
    5: "我主要看英文，跳过此题",
}
Q12 = {
    1: "不愿付",
    2: "¥20以下/月",
    3: "¥20–40/月",
    4: "¥40–70/月",
    5: "¥70–100/月",
    6: "¥100+/月",
    7: "更想买年费",
}
Q13 = {1: "手机", 2: "电脑", 3: "都用"}
Q15 = {1: "几乎每天", 2: "一周几次", 3: "偶尔", 4: "很少或没有", 5: "不确定"}
Q16 = {
    1: "当作主要依据",
    2: "当作参考之一",
    3: "有机会会去核对原文",
    4: "基本忽略",
    5: "没见过这类内容",
}
Q18 = {
    1: "语言",
    2: "付费墙",
    3: "网络访问",
    4: "不知道好的",
    5: "更信本地创作者或翻译帖",
    6: "无障碍",
}
Q19 = {1: "很感兴趣", 2: "有点兴趣", 3: "一般", 4: "不感兴趣"}
Q20 = {1: "经常", 2: "一两回", 3: "听说过", 4: "没有"}
Q24 = {
    1: "关注订阅",
    2: "考虑付费会员",
    3: "只用免费",
    4: "不感兴趣",
    5: "先看质量",
}
D1 = {1: "<18", 2: "18–24", 3: "25–34", 4: "35–44", 5: "45+", 6: "不愿说"}
D2 = {1: "男", 2: "女", 3: "其他/不愿说"}

Q3_LABELS = ["小联赛中胜出", "冲总榜", "看内容社区", "学球", "随便玩玩"]
Q5_LABELS = [
    "官方FPL",
    "海外FPL网站或会员工具",
    "海外FPL内容创作者",
    "中文FPL创作者·公众号·KOL",
    "本地翻译/改写海外内容帖",
    "微信群·朋友圈",
    "小红书",
    "抖音·B站等短视频",
    "Reddit·Discord·Telegram",
    "只问朋友",
    "AI聊天工具",
    "其他",
]
Q8_LABELS = [
    "直接看英文原文",
    "自己用翻译工具",
    "依赖中文创作者或翻译转载帖",
    "朋友帮忙总结",
    "很少用海外内容",
]
Q10_LABELS = ["队长", "转会", "冷门", "板凳", "Chip", "涨跌价", "赛程", "定位球与出场时间"]
Q11_LABELS = [
    "xP参考",
    "预测首发",
    "转会规划",
    "赛程FDR",
    "涨跌价预测",
    "冷门与拥有率",
    "阵容评分",
    "绑定我阵容的AI分析",
    "中文每日简报",
    "只用免费",
    "AI交互",
]
Q14_LABELS = [
    "微信群",
    "朋友圈",
    "公众号",
    "小程序",
    "QQ",
    "小红书",
    "抖音B站",
    "英文平台",
    "线下朋友",
    "不聊",
]
Q17_LABELS = ["短日报", "长文分析", "短视频", "直播问答", "互动数据表", "AI问答"]
Q21_LABELS = [
    "AI聊天",
    "Insights/洞察",
    "季前赛结果",
    "阵容搭建",
    "赛程",
    "微信日报",
    "新闻",
    "未用过",
    "其他",
]
Q25_LABELS = [
    "明确原文署名",
    "高质量翻译",
    "分析深度不输海外原文",
    "本地社群管理",
    "合理中国定价",
    "无博彩包装",
    "稳定的每周更新",
]

# Column indices (1-based) from headers
COL = {
    "time": 2,
    "duration": 3,
    "source": 4,
    "ip": 6,
    "S1": 8,
    "S2": 9,
    "S3": 10,
    "Q1": 11,
    "Q2": 12,
    "Q3_start": 13,
    "Q4": 18,
    "Q5_start": 19,
    "Q6_start": 31,
    "Q7": 43,
    "Q8_start": 44,
    "Q9": 49,
    "Q10_start": 50,
    "Q11_start": 58,
    "Q12": 69,
    "Q13": 70,
    "Q14_start": 71,
    "Q15": 81,
    "Q16": 82,
    "Q17_start": 83,
    "Q18": 89,
    "Q19": 90,
    "Q20": 91,
    "Q21_start": 92,
    "Q22": 101,
    "Q23": 102,
    "Q24": 103,
    "Q25_start": 104,
    "Q26": 111,
    "D1": 112,
    "D2": 113,
}


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def pct(n, d):
    return round(100.0 * n / d, 1) if d else 0.0


def multi_counts(rows, start_col, labels):
    c = Counter()
    for r in rows:
        for i, lab in enumerate(labels):
            if to_int(r[start_col + i - 1]) == 1:  # 0-index row list vs 1-index col
                c[lab] += 1
    return c


def single_counts(rows, col, mapping):
    c = Counter()
    for r in rows:
        v = to_int(r[col - 1])
        if v is None:
            continue
        c[mapping.get(v, f"unknown:{v}")] += 1
    return c


def is_chinaish(s2):
    return s2 in (1, 2, 3)  # 大陆 / 港澳台 / 新马


def is_cn_lang(s3):
    return s3 in (2, 3)  # 仅中文 / 双语


def brand_hits(text: str) -> list[str]:
    if not text:
        return []
    t = text.lower()
    patterns = [
        ("fantasy football scout", "Fantasy Football Scout"),
        ("ffscout", "FFScout"),
        ("scout", "Scout (generic)"),
        ("livefpl", "LiveFPL"),
        ("fplreview", "FPL Review"),
        ("fantasy football fix", "Fantasy Football Fix"),
        ("fplfix", "FPL Fix"),
        ("letstalk", "Let's Talk FPL"),
        ("fpl harry", "FPL Harry"),
        ("fpl general", "FPL General"),
        ("fpl focal", "FPL Focal"),
        ("faleague", "FALEAGUE"),
        ("official", "Official FPL"),
        ("官方", "Official FPL"),
    ]
    found = []
    for key, label in patterns:
        if key in t or key in text:
            found.append(label)
    return found


def main():
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])

    n = len(rows)
    china = [r for r in rows if is_chinaish(to_int(r[COL["S2"] - 1]))]
    n_cn = len(china)
    focus = china if n_cn >= max(20, int(0.5 * n)) else rows
    focus_label = "Greater China + SG/MY" if focus is china else "All respondents"
    n_f = len(focus)

    # Sample composition
    s1 = single_counts(focus, COL["S1"], S1)
    s2_all = single_counts(rows, COL["S2"], S2)
    s3 = single_counts(focus, COL["S3"], S3)
    q1 = single_counts(focus, COL["Q1"], Q1)
    q2 = single_counts(focus, COL["Q2"], Q2)
    d1 = single_counts(focus, COL["D1"], D1)
    d2 = single_counts(focus, COL["D2"], D2)
    source = Counter(str(r[COL["source"] - 1] or "未知") for r in rows)

    q3 = multi_counts(focus, COL["Q3_start"], Q3_LABELS)
    q4 = single_counts(focus, COL["Q4"], Q4)
    q5 = multi_counts(focus, COL["Q5_start"], Q5_LABELS)
    q6 = multi_counts(focus, COL["Q6_start"], Q5_LABELS)
    q8 = multi_counts(focus, COL["Q8_start"], Q8_LABELS)
    q9 = single_counts(focus, COL["Q9"], Q9)
    q10 = multi_counts(focus, COL["Q10_start"], Q10_LABELS)
    q11 = multi_counts(focus, COL["Q11_start"], Q11_LABELS)
    q12 = single_counts(focus, COL["Q12"], Q12)
    q13 = single_counts(focus, COL["Q13"], Q13)
    q14 = multi_counts(focus, COL["Q14_start"], Q14_LABELS)
    q15 = single_counts(focus, COL["Q15"], Q15)
    q16 = single_counts(focus, COL["Q16"], Q16)
    q17 = multi_counts(focus, COL["Q17_start"], Q17_LABELS)
    q18 = single_counts(focus, COL["Q18"], Q18)
    q19 = single_counts(focus, COL["Q19"], Q19)
    q20 = single_counts(focus, COL["Q20"], Q20)
    q21 = multi_counts(focus, COL["Q21_start"], Q21_LABELS)
    q24 = single_counts(focus, COL["Q24"], Q24)
    q25 = multi_counts(focus, COL["Q25_start"], Q25_LABELS)

    # Derived KPIs for deck
    overseas_site = q5.get("海外FPL网站或会员工具", 0)
    overseas_creator = q5.get("海外FPL内容创作者", 0)
    local_translate = q5.get("本地翻译/改写海外内容帖", 0)
    wechat_src = q5.get("微信群·朋友圈", 0)
    cn_kol = q5.get("中文FPL创作者·公众号·KOL", 0)
    xhs = q5.get("小红书", 0)

    use_overseas_direct = q8.get("直接看英文原文", 0)
    use_translate_tool = q8.get("自己用翻译工具", 0)
    use_cn_repost = q8.get("依赖中文创作者或翻译转载帖", 0)
    use_friends = q8.get("朋友帮忙总结", 0)
    use_rarely = q8.get("很少用海外内容", 0)

    see_translated_often = q15.get("几乎每天", 0) + q15.get("一周几次", 0)
    act_main = q16.get("当作主要依据", 0)
    act_one_input = q16.get("当作参考之一", 0)
    act_check_orig = q16.get("有机会会去核对原文", 0)

    hub_interest = q19.get("很感兴趣", 0) + q19.get("有点兴趣", 0)
    pay_any = n_f - q12.get("不愿付", 0)
    pay_20_plus = (
        q12.get("¥20–40/月", 0)
        + q12.get("¥40–70/月", 0)
        + q12.get("¥70–100/月", 0)
        + q12.get("¥100+/月", 0)
        + q12.get("更想买年费", 0)
    )
    free_only = q11.get("只用免费", 0)

    partner_follow = q24.get("关注订阅", 0)
    partner_paid = q24.get("考虑付费会员", 0)
    partner_quality = q24.get("先看质量", 0)

    wechat_group = q14.get("微信群", 0)

    # NPS among users
    nps_scores = []
    for r in focus:
        used = to_int(r[COL["Q20"] - 1])
        if used in (1, 2):
            sc = to_int(r[COL["Q22"] - 1])
            if sc is not None:
                nps_scores.append(sc)
    nps = None
    if nps_scores:
        promoters = sum(1 for s in nps_scores if s >= 9)
        detractors = sum(1 for s in nps_scores if s <= 6)
        nps = round(100 * (promoters - detractors) / len(nps_scores), 1)

    # Unaided brands
    brand_c = Counter()
    q7_raw = []
    for r in focus:
        t = r[COL["Q7"] - 1]
        if t and str(t).strip() and str(t).strip().lower() not in ("无", "没有", "无", "-", "n/a"):
            q7_raw.append(str(t).strip())
            for b in brand_hits(str(t)):
                brand_c[b] += 1

    q26_raw = [
        str(r[COL["Q26"] - 1]).strip()
        for r in focus
        if r[COL["Q26"] - 1] and str(r[COL["Q26"] - 1]).strip()
    ]
    q23_raw = [
        str(r[COL["Q23"] - 1]).strip()
        for r in focus
        if r[COL["Q23"] - 1] and str(r[COL["Q23"] - 1]).strip()
    ]

    def top_pct(counter, base=n_f, k=8):
        return [
            {"label": lab, "n": c, "pct": pct(c, base)}
            for lab, c in counter.most_common(k)
        ]

    summary = {
        "file": xlsx.name,
        "n_total": n,
        "n_focus": n_f,
        "focus_label": focus_label,
        "n_china_plus": n_cn,
        "geo_all": {k: v for k, v in s2_all.most_common()},
        "source": dict(source),
        "kpis": {
            "pct_use_overseas_site_or_creator": pct(
                len(
                    [
                        r
                        for r in focus
                        if to_int(r[COL["Q5_start"] + 1 - 1]) == 1
                        or to_int(r[COL["Q5_start"] + 2 - 1]) == 1
                    ]
                ),
                n_f,
            ),
            "pct_local_translate_posts_as_source": pct(local_translate, n_f),
            "pct_wechat_as_source": pct(wechat_src, n_f),
            "pct_cn_kol_as_source": pct(cn_kol, n_f),
            "pct_rely_cn_repost_path": pct(use_cn_repost, n_f),
            "pct_see_translated_weekly_plus": pct(see_translated_often, n_f),
            "pct_use_translated_as_input": pct(act_main + act_one_input, n_f),
            "pct_hub_interested": pct(hub_interest, n_f),
            "pct_willing_pay_any": pct(pay_any, n_f),
            "pct_willing_pay_20plus_or_annual": pct(pay_20_plus, n_f),
            "pct_partner_follow_or_paid": pct(partner_follow + partner_paid, n_f),
            "pct_partner_quality_first": pct(partner_quality, n_f),
            "pct_wechat_group_browse": pct(wechat_group, n_f),
            "faleague_nps": nps,
            "faleague_nps_n": len(nps_scores),
        },
        "distributions": {
            "S1_play": top_pct(s1),
            "S3_lang": top_pct(s3),
            "Q1_seasons": top_pct(q1),
            "Q2_rank": top_pct(q2),
            "Q3_motive": top_pct(q3),
            "Q4_hours": top_pct(q4),
            "Q5_sources": top_pct(q5, k=12),
            "Q6_top3": top_pct(q6, k=12),
            "Q8_overseas_path": top_pct(q8),
            "Q9_en_trust": top_pct(q9),
            "Q10_hardest": top_pct(q10),
            "Q11_pay_tools": top_pct(q11, k=11),
            "Q12_wtp": top_pct(q12),
            "Q13_device": top_pct(q13),
            "Q14_channels": top_pct(q14, k=10),
            "Q15_translated_freq": top_pct(q15),
            "Q16_translated_use": top_pct(q16),
            "Q17_formats": top_pct(q17),
            "Q18_barrier": top_pct(q18),
            "Q19_hub": top_pct(q19),
            "Q20_faleague": top_pct(q20),
            "Q21_useful": top_pct(q21),
            "Q24_partner": top_pct(q24),
            "Q25_trust": top_pct(q25),
            "D1_age": top_pct(d1),
            "D2_gender": top_pct(d2),
        },
        "unaided_brands": top_pct(brand_c, base=len(q7_raw) or 1, k=15),
        "q7_n_answered": len(q7_raw),
        "q7_samples": q7_raw[:25],
        "q23_samples": q23_raw[:20],
        "q26_samples": q26_raw[:25],
    }

    OUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    # Markdown for deck
    k = summary["kpis"]

    def fmt_dist(key, top=5):
        items = summary["distributions"][key][:top]
        return " · ".join(f"{i['label']} {i['pct']}%" for i in items)

    md = f"""# Survey insights for Scout deck (v2 evidence)
**Source:** Wenjuanxing export · n={n} · Focus cut: {focus_label} (n={n_f}) · China+/SG-MY geo count={n_cn}  
**Bias note:** Overseas brand names were hidden in fielding; unaided Q7 coded after the fact.

## Sample caveats (say this on the slide)
- Convenience sample via WeChat (source mix: {dict(source)}) — **not** a nationally representative China FPL census.
- Use as **directional evidence** for behaviour & willingness-to-pay, not for “China has X million FPLers.”

## Deck slide pack (copy-ready)

### Slide A — Who answered
- n={n_f} in focus cut ({focus_label})
- Playing now: {fmt_dist('S1_play', 4)}
- Language: {fmt_dist('S3_lang', 4)}
- Tenure: {fmt_dist('Q1_seasons', 4)}
- Rank band: {fmt_dist('Q2_rank', 5)}
- Device: {fmt_dist('Q13_device', 3)}
- Age (optional): {fmt_dist('D1_age', 5)}

### Slide B — How Chinese managers get FPL info (core narrative)
**Sources used regularly (Q5, multi):**  
{fmt_dist('Q5_sources', 8)}

**Most important (Q6 top-3 picks):**  
{fmt_dist('Q6_top3', 6)}

**KPIs**
- Overseas site **or** overseas creator as a regular source: **{k['pct_use_overseas_site_or_creator']}%**
- Local translated/rewritten overseas posts as a regular source: **{k['pct_local_translate_posts_as_source']}%**
- WeChat groups/Moments as a regular source: **{k['pct_wechat_as_source']}%**
- Chinese creators / 公众号 / KOL: **{k['pct_cn_kol_as_source']}%**

### Slide C — The translation bridge (your China behaviour thesis)
**How they use overseas content (Q8):** {fmt_dist('Q8_overseas_path', 5)}  
- Rely on Chinese creators / translated reposts: **{k['pct_rely_cn_repost_path']}%**

**See translated overseas tips (Q15):** {fmt_dist('Q15_translated_freq', 5)}  
- Almost daily or a few times/week: **{k['pct_see_translated_weekly_plus']}%**

**When they see those posts (Q16):** {fmt_dist('Q16_translated_use', 5)}  
- Treat as main input **or** one input among several: **{k['pct_use_translated_as_input']}%**

**English analysis dependence (Q9):** {fmt_dist('Q9_en_trust', 5)}

**Barrier to using overseas tools directly (Q18):** {fmt_dist('Q18_barrier', 6)}

### Slide D — Jobs to be done & monetisation
**Hardest GW decisions (Q10):** {fmt_dist('Q10_hardest', 6)}

**Tools they’d pay for (Q11):** {fmt_dist('Q11_pay_tools', 8)}

**WTP monthly (Q12, CNY):** {fmt_dist('Q12_wtp', 7)}  
- Any paid intent (not “不愿付”): **{k['pct_willing_pay_any']}%**  
- ¥20+/mo or annual preference: **{k['pct_willing_pay_20plus_or_annual']}%**

**Preferred Chinese formats (Q17):** {fmt_dist('Q17_formats', 6)}

**Where they browse/discuss (Q14):** {fmt_dist('Q14_channels', 6)}  
- WeChat group: **{k['pct_wechat_group_browse']}%**

### Slide E — Interest in attributed China hub / overseas-brand partnership
**Trusted Chinese hub with overseas-grade analysis + attribution (Q19):** {fmt_dist('Q19_hub', 4)}  
- Interested (很感兴趣 + 有点兴趣): **{k['pct_hub_interested']}%**

**If a well-known overseas FPL research brand partnered for Chinese content/tools (Q24):** {fmt_dist('Q24_partner', 5)}  
- Follow/subscribe **or** consider paid: **{k['pct_partner_follow_or_paid']}%**  
- Quality-first: **{k['pct_partner_quality_first']}%**

**Trust conditions (Q25):** {fmt_dist('Q25_trust', 7)}

### Slide F — Product signal (FALEAGUE) — optional / secondary
**Usage (Q20):** {fmt_dist('Q20_faleague', 4)}  
**Useful parts among respondents (Q21):** {fmt_dist('Q21_useful', 6)}  
**NPS among users who tried (Q22):** {k['faleague_nps']} (n={k['faleague_nps_n']})

### Slide G — Unaided trust names (Q7) — code carefully
Answered open trust list: n={summary['q7_n_answered']}  
Coded keyword hits (not prompted): {summary['unaided_brands'][:10]}  
Samples: {summary['q7_samples'][:12]}

## One-line story for the CEO
Chinese-speaking FPLers in this WeChat-recruited sample already pull overseas signal — often via **local translation/repost paths and WeChat** — and a clear majority show interest in a **trusted, attributed Chinese hub**; monetisation exists but is price-sensitive in ¥.

## What NOT to claim
- “China FPL market size = …” from n={n}
- Brand awareness % for any single overseas brand unless based on unaided coding with disclosed n
"""

    OUT_MD.write_text(md, encoding="utf-8")
    print(json.dumps({"n": n, "n_focus": n_f, "kpis": k, "out_md": str(OUT_MD)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
